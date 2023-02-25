import streamlit as st
import streamlit_authenticator as stauth
import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from pytz import timezone
import time as ttime
from google.cloud import firestore
from google.oauth2 import service_account
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Fill, Border, Side
from openpyxl.utils import get_column_letter
import io



st.set_page_config(layout='wide')


#  This is the APP Class


class App:
    
    """
    Here is some initial setup that get's run on each refresh.
    # I could utilise session_state to store some variables instead of having them
    # rewritten eachtime.
    
    """
    key_dict = json.loads(st.secrets["cool_secret"]["textkey"])
    creds = service_account.Credentials.from_service_account_info(key_dict)
    db = firestore.Client(credentials=creds)

   

    body = ['Timesheet', 'Materials', 'Notes']
    item3 = None
    new_form = False
    p = 0
    total_hours = []
    
    

    """
    Here is the database file we read from Firestore.
    st.cache is used to not read this over and over after every interaction.
    
    """
    @st.cache_resource(show_spinner=False)
    def read_firestore(_self):
        
        
        _self.doc_ref = _self.db.collection("JFJ Joinery")
        _self.doc_ref = _self.doc_ref.stream()

    
        jobs = {doc.id:doc.to_dict() for doc in _self.doc_ref}
        return jobs


    """
    This writes the data to Firestore
    
    """
    def write_to_json(self, text, job):
        
        doc_ref = self.db.collection("JFJ Joinery").document(job)
        

        doc_ref.set(text[job])


    """
    This initialises session_states for each job
    Currently, it is putting a variable for a button in deleting times
    
    """
    def initialise_sessions(self, jobs, thing):
        if jobs != 0:
            for job in jobs:
                if job == 'Other':
                    pass
                elif job == 'config':
                    pass
                
                elif f'{job}_button' not in st.session_state:
                    st.session_state[f'{job}_button'] = 'False'

                st.session_state["Sheet_button"] = 'False'
            return
        elif jobs == 0:
            if thing == 0:
                return
            st.session_state[f'{thing}'] = 'False'

       

    """
    This makes a list of available times in 5 minute intervals.

    """
    # Can add this into st.cache. When they fix the bug
    def time_list(self):

        hours = ['00', '01', '02', '03', '04', '06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23']
        minutes = ['00','05','10','15','20','25','30','35','40','45','50','55']

        times = []

        for h in hours:
            for m in minutes:
                times.append(f'{h}:{m}')

        return times[::-1]



    """
    This function updates the maintains the Tally at the top of the screen.
    
    """
    def update_hours_tally(self, jobs, type):
        hour_list = []
        time = jobs["Other"]["Employee_Variables"][f'{st.session_state["name"]}']
        if type == 1:
            for punches in time["Total_Hours"]:
                if punches["Date"] == self.get_current_date():
                    checkin = self.convert_to_int(punches["Check-in"])
                    checkout = self.convert_to_int(punches["Check-out"])
                    if punches["Break"] == 'Yes':
                        break_ = 0.5
                    else:
                        break_ = 0
                    value = checkout - checkin - break_
                    hour_list.append(value)


            try:
                checkin_ = self.convert_to_int(time["Start"])
                checkout_ = self.convert_to_int(self.get_current_time())
                value_ = checkout_ - checkin_
                hour_list.append(value_)
            except:
                pass

        if type == 2:
            checkin = self.convert_to_int(time["Start"])
            checkout = self.convert_to_int(time["End"])
            hours = checkout - checkin
            
            

        if type == 1:      
            hours = 0
            for ele in range(0, len(hour_list)):
                hours = hours + hour_list[ele]
                hours = round(hours, 1)

        try:
            if hours > 8:
                OT = hours - 8


            else:

                OT = 0
        except:
            hours = 0
            OT = 0

        return OT, hours


    """
    This fuction retruns the current date in Day / Month / Year format
    
    """
    def get_current_date(self):
        date = datetime.today()
        date = date.strftime("%d/%m/%Y")
        return date


    """
    This function converts st.date_input output into the above format.
    
    """
    def convert_date(self, date):
        date = str(date).split('/')
        date = f"{date[2]}/{date[1]}/{date[0]}"
        return date


    """
    This function returns the current time in text format EG '00:00'
    
    """
    def get_current_time(self):
        tz = timezone('Australia/Victoria')
        time = datetime.now(tz)
        time = str(time).split(' ')
        time = time[1].split(':')
        
        if len(time[0]) == 1:
            
            time[0] = f'0{time[0]}'
        time = f'{time[0]}:{time[1]}'
        
        return time


    """
    This function rounds the hours in the Tally to a single decimal value
    
    """
    def round_hours(self, hours):
        hours = round(float(hours),1)
        return hours


    """
    This function converts '00:00' time into a int decimal value
    
    """
    def convert_to_int(self, value):
        h, m = value.split(":")
        m = int(m) / 60
        value = int(h) + m
        
        return value

    
    """
    This function converts int / decimal value into string format '00:00'
    
    """
    def convert_to_time(self, value):
        h, m = str(value).split('.')
        m = f"0.{m}"
        m = float(m) * 60
        m = round(m)
        if len(str(m)) == 1:
            m = f'0{m}'
        
        if len(str(h)) == 1:
            h = f'0{h}'

        time_ = f'{h}:{m}'
        
        return time_
                          


    """
    This function rounds the time to the nearest value specified by x.
    EG: x=15, will round to the nearest 15minute mark.
    
    """
    def round_time(self, time, x):
        h, m = time.split(":")
        m = self.myround(int(m), x)
        h = int(h)
        if m == 60:
            m = 0
            h += 1
        if m == 0:
            m = '00'
        else:
            m = str(m)
        if h < 10: 
            h = f'0{h}'
        if len(str(m)) == 1:
            m = f'0{m}'
        time = f'{h}:{m}'
      
        return time


    """
    A function used by the above function.
    
    """
    def myround(self, x, y):
        return y * round(x/y)


    """
    This function Authenticates the session and is used for website cookies
    and storing login credentials
    
    """
    def auth_read(self, config):


        authenticator = stauth.Authenticate(
                    config['credentials'],
                    config['cookie']['name'],
                    config['cookie']['key'],
                    config['cookie']['expiry_days'],
                    config['preauthorized']
                )

        return authenticator


    
    """
    This retrieves a list of employee's registered
    
    """
    def get_employee_list(self, jobs, job):
        employees = []
        for keys, items in jobs[job]["Employee_Variables"].items():
            employees.append(keys)
        employees.sort()
        return employees

    
    """
    This function creates a new user within the config file
    
    
    """
    def update_credentials(self, first, last, email, passw, jobs):
        hashed_password = stauth.Hasher([passw]).generate()
        new = {
                                "email": email,
                                "name": f'{first} {last}',
                                "password": hashed_password[0]
        }
        jobs['config']['credentials']['usernames'][first.lower()] = new
        self.write_to_json(jobs, 'config')


    """
    This funciton is FIXME: This will return hours worked on various jobs
    as well as total hours worked for the day/ period.
    
    """
    def shift_search(self, jobs, employee, date_):
        date_ = str(date_).split('-')
        date_ = f'{date_[2]}/{date_[1]}/{date_[0]}'
        total_hours = []
        job_hours = []
        for job in jobs:
            if job == 'Other':
                
                if len(employee) == 1:
                    for shift in jobs[job]['Employee_Variables'][employee[0]]["Total_Hours"]:
                        if shift['Date'] == date_:
                            total_hours.append([[shift['Date'],shift['Check-in'],shift['Check-out'],8]])
                    

    """
    FIXME: This is a function to confirm the delete of shift hours.
    Still massively under progress.
    
    """
    def confirm_delete(self):
        st.write("Are you sure you want to delete?")
        yes = st.button("YES")
        no = st.button("NO")
        if yes:
            self.choice = 0
        elif no:
            self.choice = 1


    """
    This is a switch for job buttons when editing the shift times.
    It allows the code to click on a single edit button, 
        without opening all edit field for all jobs.
        
    """
    def edit_button(self, job):
        
        if st.session_state[f"{job}_button"] == 'False':
            st.session_state[f"{job}_button"] = 'True'
        elif st.session_state[f"{job}_button"] == 'True':
            st.session_state[f"{job}_button"] = 'False'
        return


    """
    This fetches a list of jobs, this can potentially be placed with an
        st.cache_data. And then build other widgets to use this data.

    """
    def job_list_fetch(self, jobs):
        self.j_list =[""]
        
        for job in jobs:
            if job == 'Other':
                pass
            elif job == 'config':
                pass
            else:
                
                self.j_list.append(job)
        return self.j_list


    """
    TEST
    
    """
    def timesheet_tool(self, datestart, dateend, employees, jobs):
        daterange = pd.date_range(datestart, dateend).strftime("%d/%m/%Y")
        job_l = []
        job_l_array = []
        shed_l = []
        shed_l_array = []

        for name in employees:
            for date in daterange:
                for job in jobs:
                    if job == 'Other':
                        for shed_hours in jobs[job]["Employee_Variables"][name]["Shed_Hours"]:
                            if shed_hours["Date"] == str(date):
                                shed_l.append(shed_hours)



                            

                    elif job == 'config':
                        pass
                    else:
                        
                        for shift in jobs[job]["Timesheets"]:

                            if shift['Name'] == name:
                            
                                if shift['Date'] == str(date):
                                
                                    job_l.append(shift)

                jobo = job_l.copy()
                job_l_array.append(jobo)
                job_l.clear()
            shedo = shed_l.copy()
            shed_l_array.append(shedo)
            shed_l.clear()



            
        


        return job_l_array, shed_l_array

    
    def create_timesheet_spreadsheet(self, array, array2):
        workbook = Workbook()
        sheet = workbook.active
        values = self.spread_sheet_value_list()
        corrected_shed_hours = self.evaluate_hours(array2)
        print('$$$$$$$$$$')
        print(corrected_shed_hours)
        
        jobs = []
        dates = []
        names = []
        items_ = []
        things_ = []
        name_r = 3
        abc = 1
        dates_ = 1
        write = {}
        colors = ['00CCFFFF', '00CCCCFF', '00CCFFCC', '00FFFFCC', '00FFFF00', '00FFCC99']
        for items in array:
            for item in items:
                if item['Job'] not in jobs:
                    jobs.append(item['Job'])
                if item['Date'] not in dates:
                    dates.append(item['Date'])
                if item['Name'] not in names:
                    names.append(item['Name'])
                items_.append(item)
        spacing_ = len(jobs) + 1

        for things in corrected_shed_hours:
            for thing in things:
                if thing['Date'] not in dates:
                    dates.append(thing['Date'])
                if thing['Name'] not in names:
                    names.append(thing['Name'])
                things_.append(thing)
        dates.sort()

        t_font = Font(bold=True)
        d_font = Font(bold=True, size=13)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin')
                             )


        for date in dates:
            
            sheet.merge_cells(f'{values[dates_]}1:{values[(dates_ +spacing_)]}1')
            top_left_cell = sheet[f'{values[dates_]}1']
            top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
            DATE = sheet[f'{values[dates_]}1']
            DATE.font = d_font
            DATE.value = date
            DATE.fill = PatternFill('solid', '0099CCFF')
            DATE.border = thin_border
            

            dates_ += spacing_ + 1
            

        jobs.sort()    
        didx = 1
        spacing_ = len(jobs) + 2

        for date in dates:
            for job in jobs:
            
                JOB = sheet[f'{values[abc]}2']
                JOB.font = t_font
                JOB.value = job
                JOB.border = thin_border
                abc += 1

            

            SHED = sheet[f'{values[abc]}2']
            SHED.font = t_font
            SHED.value = 'Shed'
            SHED.border = thin_border
            abc += 1
            TOTAL = sheet[f'{values[abc]}2']
            TOTAL.font = t_font
            TOTAL.value = 'TOTAL'
            TOTAL.border = thin_border
            abc += 1

            nidx = 3
            endc = didx * spacing_
            reverseE = endc - 1
            reverseS = endc - (len(jobs) + 1)
            for name in names:

                total = sheet[f'{values[(endc)]}{nidx}']
                total.value = f'=sum({values[reverseS]}{nidx}:{values[reverseE]}{nidx})'
                total.border = thin_border
                nidx += 1
            
            
            base_total = sheet[f'{values[(endc)]}{name_r}']
            base_total.value = f'=sum({values[reverseS]}{name_r}:{values[reverseE]}{name_r})'
            base_total.border = thin_border
            didx += 1

            

            
        GRANDTOTAL = sheet[f'{values[abc]}2']
        
        GRANDTOTAL.font = t_font
        GRANDTOTAL.value = 'GRAND TOTAL'
        GRANDTOTAL.border = thin_border
        
        endCol = values[abc]
        names.sort()

        for name in names:
            NAME = sheet[f'A{name_r}']
            NAME.value = name
            NAME.border = thin_border
            name_r += 1
        TOTAL2 = sheet[f'A{name_r}']
        TOTAL2.font = t_font
        TOTAL2.value = 'TOTAL'
        TOTAL2.border = thin_border
        spacing_ = len(jobs) + 2
        

        for thing_ in things_:
            if thing_['Name'] in names:
                nidx = names.index(thing_['Name']) + 3

                if thing_['Date'] in dates:
                    didx = dates.index(thing_['Date'])
                    xxx = didx * spacing_

                    sidx = len(jobs) + 1

                    sheet[f'{values[sidx + xxx]}{nidx}'] = thing_['Hours']
        for item_ in items_:
            if item_['Name'] in names:
                nidx = names.index(item_['Name']) + 3

                if item_['Date'] in dates:
                    didx = dates.index(item_['Date'])
                    xxx = didx * spacing_
                    endci = didx + 1
                    endc = spacing_ * endci
                    reverseE = endc - 1
                    reverseS = endc - (len(jobs) + 1)
                    sheet[f'{values[(endc)]}{nidx}'] = f'=sum({values[reverseS]}{nidx}:{values[reverseE]}{nidx})'
                    sheet[f'{values[(endc)]}{name_r}'] = f'=sum({values[reverseS]}{name_r}:{values[reverseE]}{name_r})'

                    ve = reverseE - reverseS + 1
                    
                    while ve > 0:
                        sheet[f'{values[(endc-ve)]}{name_r}'] = f'=sum({values[(endc-ve)]}{name_r-1}:{values[(endc-ve)]}{3})'
                        n = len(names)
                        
                        while n > -1:
                            colored_cell = sheet.cell(column=endc-n, row=name_r - ve)
                            colored_cell.fill = PatternFill('solid', fgColor=colors[n])
                            colored_cell.border = thin_border
                            n -= 1
                        ve -= 1
                        

                    if item_['Job'] in jobs:
                        
                        jidx = jobs.index(item_['Job']) + 2

                        if f'{values[(jidx+xxx)-1]}{nidx}' in write:
                            combine = int(write[f'{values[(jidx+xxx)-1]}{nidx}']) + int(item_['Total'])
                            write[f'{values[(jidx+xxx)-1]}{nidx}'] = combine
                            
                        else:
                            write[f'{values[(jidx+xxx)-1]}{nidx}'] = item_['Total']
        
                        

                        sheet[f'{values[(jidx+xxx)-1]}{nidx}'] = write[f'{values[(jidx+xxx)-1]}{nidx}']
                gt = (len(dates) * (len(jobs) + 2)) + 1
                ts = gt - 1
                ti = ts / len(dates)
                tl = []
                tll = []
                while ts > 0:
                    tl.append(ts)
                    ts = ts - ti
                for v in tl:
                    
                    tll.append(values[int(v)])
                mystring = ''.join([str(item + f'{nidx},') for item in tll])
                sheet[f'{values[gt]}{nidx}'] = f'=sum({mystring})'
                sheet[f'{values[gt]}{name_r}'] = f'=sum({values[gt]}{name_r-1}:{values[gt]}{3})'
            
   
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
                else:
                    column_widths += [len(str(cell))]

        for i, column_width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width
        freeze = sheet['B1']
        sheet.freeze_panes = freeze
        return workbook

    
    def evaluate_hours(self, array):
        print('------------------')
        print(array)
        new_l = []
        new_array = []
        
        for items in array:
            for shift in items:

                for i in new_l:
                    if i['Date'] == shift['Date']:
                        
                        i['Hours'] = shift['Hours'] + i['Hours']
                        print('hi')
                    else:
                        new = {'Hours': shift['Hours'],
                               'Name': shift['Name'],
                               'Date': shift['Date'],
                               'OT': shift['OT']}
                        new_l.append(new)

                if new_l == []:
                    new = {'Hours': shift['Hours'],
                            'Name': shift['Name'],
                            'Date': shift['Date'],
                            'OT': shift['OT']}
                    new_l.append(new)

                    


            newo = new_l.copy()
            new_array.append(newo)
            new_l.clear()
        
        
        

        return new_array
                



    def spread_sheet_value_list(self):
        ssvl = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
        ssvl_ = ssvl.copy()
        ssvl__ = ssvl.copy()
        extended_ssvl = []
        for item in ssvl:
            for i in ssvl_:
                extended_ssvl.append(f'{item}{i}')

        ssvl__.extend(extended_ssvl)
        return ssvl__



    def calc_hours(self, in_, out_):
        ts = in_
        te = out_
        ts = self.convert_to_int(ts)
        te = self.convert_to_int(te)
        tf = te - ts
        tf = round(tf, 1)

        return  tf
        

        
    """
    This is the main function. This is where the apps code and commands come from.
    
    """
    def main(self):
        
        # This reads the database
        jobs = self.read_firestore()
        
        # This sets up the Authenticator for each user.
        authenticator = self.auth_read(jobs['config'])
        
        # Sets variables for the login page.
        name, authentication_status, username = authenticator.login('Login', 'main')
        
        
        # Determines what happens based on status. None means no one is logged in.
        if authentication_status == None:
            if st.sidebar.button('Register'):
                print('hi')
            if st.sidebar.button('Forgot Password'):
                print('bye')

        # If you login, status = True and then starts the main script.
        if st.session_state["authentication_status"]:

            self.initialise_sessions(jobs, 0)

            # This is going to determine all the widgets available in the sidebar.
            with st.sidebar:

                # Logout widget
                authenticator.logout('Logout', 'main')

                # These widgets will only be available to those with permissions.
                # EG, boss, administrators.
                if jobs["Other"]["Employee_Variables"][f"{st.session_state['name']}"]["Auth"] == ("Admin"):
                    st.subheader('')
                    st.subheader(':blue[User Tools]')

                    # This widget will allow you to add new employees to be able to
                    #   Login.
                    with st.sidebar.expander('Add Employee', expanded=False):

                        with st.form(key='test9', clear_on_submit=True):
                            
                            fname_ = st.text_input('First Name', key='new_name', value='')
                            lname_ = st.text_input('Last Name', key='new_last', value='')
                            email_ = st.text_input('Email', key='email', value='@')
                            passw_ = st.text_input('Temp Pass', key='pass', value='')
                            autho_ = st.selectbox('User Type', ['Admin','Standard'])


                            submitted1 = st.form_submit_button('Submit')

                            if submitted1:
                                
                                new = {"Auth": autho_,
                                        "Break": "False",
                                        "Check-in": "False",
                                        "End": "",
                                        "First_Login": True,
                                        "Start": "",
                                        "Total_Hours": [],
                                        "Shed_Hours": [],
                                        "Form": "False"}
                                jobs["Other"]["Employee_Variables"][f'{fname_} {lname_}'] = new

                                self.update_credentials(fname_, lname_, email_, passw_, jobs)                                
                                self.write_to_json(jobs, "Other")

                    st.subheader('')
                    st.subheader(':blue[Job Tools]')

                    # This Widget will allow you to create a blank job.
                    with st.sidebar.expander('Create New Job', expanded=False):

                        with st.form(key='test11', clear_on_submit=True):
                            job_title = st.text_input('Job Title')
                            create_ = st.form_submit_button(label="Create Job")

                            if create_:
                                if job_title in jobs:
                                    st.error('Job name already exists')
                                elif job_title not in jobs:

                                    new = {
                                                "Notes": {},
                                                "Materials": [],
                                                "Status": "Current",
                                                "Timesheets": []
                                    }
                                    jobs[job_title] = new

                                    self.initialise_sessions(jobs, 0)
                                
                                    self.write_to_json(jobs, job_title)
                                    
                    # This Widget will allow you to delete an existing job.
                    with st.sidebar.expander('Delete a Job', expanded=False):

                        with st.form(key='test13', clear_on_submit=False):
                            
                            j_list = self.job_list_fetch(jobs)

                            job_title2 = st.selectbox('Job Title', options=j_list, index=0, key=f'Job Title{2}')
                            j_select2 = st.form_submit_button("Delete Job")
                       

                            if j_select2:
                                
                                self.db.collection("JFJ Joinery").document(str(job_title2)).delete()
                                
                                del jobs[job_title2]
                                j_list = self.job_list_fetch(jobs)
                                st.experimental_rerun()
                                
                    # This Widget will allow you to move a job from 'Current' to 'Complete'
                    with st.sidebar.expander('Edit Job Status', expanded=False):

                        with st.form(key='test12', clear_on_submit=True):

                            
                            j_select = st.selectbox("Select Job", options=j_list, index=0)
                            s_select = st.selectbox("Status", options=['', "Current", "Complete"],index=0)
                            select_ = st.form_submit_button("Change Status")

                            if select_:
                                jobs[j_select]["Status"] = s_select
                                self.write_to_json(jobs, j_select)

                    st.subheader('')
                    st.subheader(':blue[Material Tools]')

                    # This Widget will allow you to import a list of materials
                    #   from a txt file.
                    with st.sidebar.expander('Import List of Materials', expanded=False):
                        
                        with st.form("test", clear_on_submit=True):
                            uploaded_file = st.file_uploader("Choose a file")
                        
                            if uploaded_file is not None:
                                bytes_data = uploaded_file.getvalue()

                                stringio = StringIO(uploaded_file.getvalue().decode("utf-8"))

                                string_data = stringio.read()
                                string_data = string_data.split('\n')
                                for item in string_data:
                                    if item in jobs["Other"]["Material List"]:
                                        pass
                                    else:
                                        jobs["Other"]["Material List"].append(item)
                                self.write_to_json(jobs, "Other")

                            finish_upload = st.form_submit_button('Finsh Import')

                    # This Widget will allow you to download the list to your PC
                    with st.sidebar.expander('Export List of Materials', expanded=False):

                        mlist = ''
                        for items in jobs["Other"]["Material List"]:
                            if items[-1] == ' ':
                                mlist += items[:-1] + '\n'
                            else:
                                mlist += items + '\n'
                        mlist = mlist[:-1]
                        
                        st.download_button('Download', data=mlist, file_name='Materials.txt')
                                
            
                    st.subheader('')
                    st.subheader(':blue[Timesheet Tools]')

                    # This Widget will allow you to search Employee Timesheets.
                    #   FIXME: This is still a WIP.
                    with st.sidebar.expander('Search Timesheets', expanded=False):
                        
                        with st.form(key='test10', clear_on_submit=True):
                            date_s = st.date_input('Date Start',key='date_s')
                            date_e = st.date_input('Date End', key='date_e')
                            emplo_ = st.multiselect('Employee', self.get_employee_list(jobs, 'Other'))
                            search_ = st.form_submit_button(label='Search Timesheet')
                            
                            if search_:
                                result = self.timesheet_tool(date_s, date_e, emplo_, jobs)
                                workbook = self.create_timesheet_spreadsheet(result[0],result[1])
                                vbook = io.BytesIO()
                                workbook.save(vbook)
                                st.session_state["Sheet_button"] = 'True'
                                                             

                                
                        if st.session_state["Sheet_button"] == 'True':            
                            down = st.download_button("Download Spreadsheet", data=vbook, mime='xlsx', file_name='test.xlsx')
                            if down:
                                st.session_state["Sheet_button"] = 'False'
                                


                         

            # This will print Company name to the top of the screen
            st.title(f':blue[{jobs["config"]["Company"]}]')

            # This will print users name onto the screen.
            st.write(f'Welcome *{st.session_state["name"]}*')

            # This creates a starter container.. Not sure if really needed.
            c = st.container()

            hours = self.update_hours_tally(jobs, 1)
            tally = pd.DataFrame(data=[[self.get_current_date(), hours[1], hours[0]]], columns=["Date", "Hours Today","Overtime"])
            
            with c:
                col1, col2 = st.columns(2)

                with col1:
                    c.dataframe(tally, width=500)

                with col2:
                    check_in = jobs["Other"]["Employee_Variables"][f"{st.session_state['name']}"]
                    if check_in["Check-in"] == "False":
                        dbut = c.button('Start Day')
                        if dbut:
                            check_in["Check-in"] = "True"
                            check_in["Start"] = self.round_time(self.get_current_time(), 15)
                            self.write_to_json(jobs, "Other")
                            
                            st.experimental_rerun()

                    elif check_in["Check-in"] == "True":
                        dbut = c.button('End Day')
                        if dbut:
                            for job in jobs:
                                if job == "Other":
                                    pass
                                elif job == 'config':
                                    pass
                                else:
                                    for shift in jobs[job]["Timesheets"]:
                                        if shift["Check-out"] == '':
                                            shift["Check-out"] = self.round_time(self.get_current_time(), 15)
                                            self.write_to_json(jobs, job)
                            self.new_form = True
                            time = self.get_current_time()
                            check_in["End"] = self.round_time(self.get_current_time(), 15)
                            check_in["Form"] = "Form1"
                    
                            self.write_to_json(jobs, "Other")
                            
                            st.experimental_rerun()
            new_form = jobs["Other"]["Employee_Variables"][f"{st.session_state['name']}"]
            if new_form["Form"] == "Form1":
                
                with st.form("EOD", clear_on_submit=False):
                    
                    hours = self.update_hours_tally(jobs, 2)
                    total = hours[0] + hours[1]
                    st.header(':blue[Break Review]')
                    st.write(f':blue[You have worked] {hours[1]} :blue[hours this shift!]')
                    st.write(":blue[Did you take a Lunch Break today?]")
                    submitted = st.form_submit_button('Yes')
                    submitted2 = st.form_submit_button('No')
                    
                    if submitted:
                        new_form["Form"] = "Form2"
                        new_form["Break"] = "Yes"

                        self.write_to_json(jobs, "Other")
                        
                        st.experimental_rerun()

                    elif submitted2:
                        new_form["Form"] = "Form2"
                        new_form["Break"] = "No"

                        self.write_to_json(jobs, "Other")
                        
                        st.experimental_rerun()


            elif new_form["Form"] == "Form2":
                time_list = self.time_list()
                x = 0

                with st.form("Confirm Hours"):
                    st.header(':blue[Review Job Hours]')
                    st.markdown(':blue[Please review your submitted hours below. Edit as required.]')
                    for job in jobs:
                        if job == 'Other':
                            pass
                        elif job == "config":
                            pass
                        else:
                            b = 1
                            for shift in jobs[job]["Timesheets"]:
                                if shift['Date'] == self.get_current_date():
                                    
                                    with st.expander(job, expanded=True):
                                        s_ = time_list.index(shift['Check-in'])
                                        try:
                                            e_ = time_list.index(shift['Check-out'])
                                        except:
                                            e_ = time_list.index("00:00")
                                        
                                        new_s = st.selectbox('Start Time', options=time_list, index=s_, key=f'form2{x}')
                                        x += 1
                                        new_e = st.selectbox('End Time', options=time_list, index=e_,key=f'form2{x}')
                                        x += 1
                                        delete = st.checkbox('Delete', key=f"form2{x}")
                                        x += 1
                                        st.caption('Having the box checked when submitting, will delete your hours from this job')
                                        x += 1
                                        
                                        job_submit = st.form_submit_button(f'Edit {job} {b}')
                                        b += 1
                                        if job_submit:
                                            s_i =jobs [job]["Timesheets"].index(shift)
                                            if delete:
                                                
                                                self.db.collection("JFJ Joinery").document(job).update({
                                                                "Timesheets": firestore.ArrayRemove([jobs[job]["Timesheets"][s_i]])
                                                })
                                                jobs[job]["Timesheets"].pop(s_i)
                                                st.experimental_rerun()

                                            else:
                                                jobs[job]["Timesheets"][s_i]["Check-in"] = new_s
                                                jobs[job]["Timesheets"][s_i]["Check-out"] = new_e
                                                self.write_to_json(jobs, job)
                                                st.experimental_rerun()
                                        
                                            

                                        

                    submitted3 = st.form_submit_button('Submit')
                if submitted3:
                    new_form["Form"] = 'Form3'
                    self.write_to_json(jobs, 'Other')
                    st.experimental_rerun()
                    
                        
                            
            elif new_form["Form"] == "Form3":
                with st.form("Adjust Hours"):
                    st.header(":blue[Adjust Total Hours]")
                    st.write(":blue[You can adjust hours here. Otherwise, please]")
                    st.write(":blue[provide a short decscription and click Finish to complete.]")
                    
                    num = st.number_input(label='p', label_visibility='collapsed', key='num')
                    desc = st.text_area(label='p', label_visibility='collapsed', key='desc')
                    submitted = st.form_submit_button("Finish")

                        
                        

                    if submitted:
                        
                        value = self.convert_to_int(new_form["End"])
                        
                        value = value + num
                        
                        value = self.convert_to_time(value)
                        


                        new = {"Date": f"{self.get_current_date()}",
                            "Check-in": f"{new_form['Start']}",
                            "Check-out": f"{value}",
                            "Break": f"{new_form['Break']}",
                            "Amend": f"{num}",
                            "Description": f"{desc}"}

                        new_form["Total_Hours"].append(new)
                        

                        new_form["Start"] = ''
                        new_form["End"] = ''
                        new_form['Break'] = 'False'                                                   
                        new_form["Check-in"] = 'False'
                        new_form["Form"] = 'False'
                        self.write_to_json(jobs, "Other")
                        hours = self.update_hours_tally(jobs, 1)
                        new_form["Shed_Hours"].append({"Name": st.session_state['name'],
                                                       "Date": f'{self.get_current_date()}',
                                                       "Hours": hours[1],
                                                       "OT": hours[0]})
                        self.write_to_json(jobs, "Other")
                        st.experimental_rerun()
                        

                    

                        
            else:
                pass
            x = 0
            st.header("")
            st.header("")
            current, complete = st.tabs(['Current', 'Complete'])


            with current:
                for job in jobs:
                    if job == "Other":
                        pass
                    elif job == 'config':
                        pass
                    else:
                        
                        
                        if jobs[job]["Status"] == "Current":
                            with st.expander(job):
                                t, m, n = st.tabs(self.body)

                                with t:

                                    with st.form(f"My Form {x}", clear_on_submit=False):
                                        col1, col2, col3, col0 = st.columns(4)

                                        with col1:

                                            item1 = st.text_input('Name', key=x, value=st.session_state["name"], disabled=True)
                                            
                                            x += 1

                                        with col2:
                                            item2 = st.date_input('Date', key=x)
                                            x += 1

                                        with col3:
                                            t_list = self.time_list()
                                            st_ = t_list.index(str(self.round_time(self.get_current_time(), 5)))
                                            item3 = st.selectbox('Time', options=self.time_list(), key=x, index=st_)
                                            x += 1


                                        with col0:
                                            pass

                                        submitted = st.form_submit_button('Submit')



                                    if submitted:
                                        y = 0
                                        
                                        
                                        test = True

                                        while test:

                                            try:
                                                if jobs[job]["Timesheets"] == []:
                                                    new = {"Name": f"{st.session_state['name']}",
                                                        "Date": f"{item2.strftime('%d/%m/%Y')}",
                                                        "Check-in": f"{str(item3)}",
                                                        "Check-out": "",
                                                        "Total":"",
                                                        "Job":f"{job}"}
                                                    jobs[job]["Timesheets"].append(new) 
                                                    self.write_to_json(jobs, job)
                                                    test = False
                                        
                                                elif jobs[job]["Timesheets"][y]["Name"] == f"{st.session_state['name']}":
                                                    if jobs[job]["Timesheets"][y]["Date"] == f'{item2.strftime("%d/%m/%Y")}':
                                                        if jobs[job]["Timesheets"][y]["Check-in"] == '':
                                                            jobs[job]["Timesheets"][y]["Check-in"] = str(item3)
                                                            self.write_to_json(jobs, job)
                                                            test = False
                                                        elif jobs[job]["Timesheets"][y]["Check-out"] == '':
                                                            jobs[job]["Timesheets"][y]["Check-out"] = str(item3)
                                                            
                                                            calc = self.calc_hours(jobs[job]["Timesheets"][y]["Check-in"],str(item3))
                                                            
                                                            jobs[job]["Timesheets"][y]["Total"] = calc
                                                            self.write_to_json(jobs, job)
                                                            
                                                            
                                                            test = False
                                                        else: 
                                                            y += 1
                                                    else:
                                                        y += 1
    
                                                else:
                                                    y += 1
                                                
                                            except:
                                                new = {"Name": f"{st.session_state['name']}",
                                                    "Date": f"{item2.strftime('%d/%m/%Y')}",
                                                    "Check-in": f"{str(item3)}",
                                                    "Check-out": "",
                                                    "Total": "",
                                                    "Job": f"{job}"}
                                                jobs[job]["Timesheets"].append(new) 
                                                self.write_to_json(jobs, job)
                                                test = False

                                        st.experimental_rerun()
                                                


                                    
                                    st.button('Edit Times', key=x, on_click=self.edit_button, args=(job,))
                                    x += 1
                                    if st.session_state[f"{job}_button"] == 'True':
                                        with st.form(f'{job}_edit_form'):
                                            
                                            times_list = []
                                            i = 0
                                            for items in jobs[job]["Timesheets"]:
                                                times_list.append(i)
                                                i += 1
                                            col1, col2, col3 = st.columns(3)
                                            with col1:
                                                job_ = st.selectbox('Select Timesheet Entry', options=times_list)
                                                testo = st.form_submit_button('Submit')
                                            with col2:
                                                st_ = t_list.index(str(self.round_time(self.get_current_time(), 5)))
                                                newStart = st.selectbox('Select New Start', options=self.time_list(), index=st_)
                                                dele = st.form_submit_button('Delete Entry')
                                            with col3:
                                                newEnd = st.selectbox('Select New End', options=self.time_list(), index=st_)

                                            
                                            
                                            if testo:
                                                jobs[job]["Timesheets"][job_]["Check-in"] = newStart
                                                jobs[job]["Timesheets"][job_]["Check-out"] = newEnd
                                                self.write_to_json(jobs, job)
                                                
                                            elif dele:
                                                
                                                self.db.collection("JFJ Joinery").document(job).update({
                                                                        "Timesheets": firestore.ArrayRemove([jobs[job]["Timesheets"][job_]])
                                                })
                                                jobs[job]["Timesheets"].pop(job_)
                                                
                                                
                                                
                                            
                                                
                                                
                                    listo = []
                                    for items, values in jobs[job].items():
                                        t = 0

                                        try:
                                            for thing in values:
                                                new = [thing['Name'], thing['Date'], thing['Check-in'], thing['Check-out']]
                                                listo.append(new)
                                            
                                        except:
                                            pass
                                    
                                    if listo != []:
                                        df = pd.DataFrame(data=listo, columns=['Name', 'Date', 'Check-in', 'Check-out'])
                                        st.dataframe(df, width=1000)
                                    
                                    
                                        
                                    

                                with n:
                                            
                                    with st.form(f'my_form{x}', clear_on_submit=True):
                                        text = st.text_area(label='Test', label_visibility='collapsed')
                                        submitted = st.form_submit_button("Submit")
                                        x += 1

                                    if submitted:
                                        now = datetime.now()
                                        dt = now.strftime("%d/%m/%Y - %H:%M:%S")
                                        jobs[job]["Notes"][f'{dt} - {st.session_state["name"]}'] = text 
                                        self.write_to_json(jobs, job)
                                        x += 1

                                    notes = []
                                    for keys, values in jobs[job]["Notes"].items():
                                        notes.append(keys)
                                    
                                    notes.sort()
                                    notes.reverse()

                                    for note in notes:

                                        st.text_area(label=note, key=x, value=jobs[job]["Notes"][note], disabled=True)
                                        x += 1
                                        

                                with m:

                                    array = []
                                    
                                    with st.form(f'My_form{x}'):
                                        material = st.selectbox('Material', options=jobs["Other"]["Material List"], key=x)
                                        x += 1
                                        quantity = st.number_input('Quantity', key=x, step=1)
                                        x += 1

                                        submitted = st.form_submit_button("Submit")
                                    
                                    if submitted:                                                   


                                        item_list = []
                                        for items in jobs[job]["Materials"]:
                                            if material in items["Material"]:
                                                
                        
                                                q = float(items["Quantity"])
                                                
                                                q += quantity
                                                
                                                items["Quantity"] = q
                                                item_list.append((items["Material"]))
                                                
                                                self.write_to_json(jobs, job)
                                                
                                            
                                        if material not in item_list:
                                            

                                            if quantity == 0:
                                                pass
                                            else:
                                                
                                                new = {
                                                    "Material": f"{material}",
                                                    "Quantity": quantity
                                                }
                                                jobs[job]["Materials"].append(new)
                                                self.write_to_json(jobs, job)
                                        
                                                    
    
                                    for selection in jobs[job]["Materials"]:
                                        
                                        
                                        array.append([selection["Material"], selection["Quantity"]])
                                        

                                        testu = False
                                    if array != []:
                                        df = pd.DataFrame(data=array, columns=["Material", "Quantity"])
                                        hide = """
                                        <style>
                                        .row_heading.level0 {display:none}
                                        .blank {display:none}
                                        </style>"""
                                        st.markdown(hide, unsafe_allow_html=True)
                                        st.dataframe(df, width=500)








            with complete:
                for job in jobs:

                    if job == 'config':
                        pass
                    if job == 'Other':
                        pass

                    try:
                        if jobs[job]["Status"] == "Complete":
                            with st.expander(job):
                                t, p, n = st.tabs(self.body)
                                with t:
                                    listo = []
                                    for items, values in jobs[job].items():
                                        t = 0

                                        try:
                                            for thing in values:
                                                new = [thing['Name'], thing['Date'], thing['Check-in'], thing['Check-out']]
                                                listo.append(new)
                                            
                                        except:
                                            pass
                                    if listo != []:
                                        df = pd.DataFrame(data=listo, columns=['Name', 'Date', 'Check-in', 'Check-out'])
                                        st.dataframe(df, width=1000)
                                with n:
                                                
                                        y = 1
                                        z = 0


                                                
                                        with st.form(f'my_form{x}', clear_on_submit=True):
                                            text = st.text_area(label='Test', label_visibility='collapsed')
                                            submitted = st.form_submit_button("Submit")
                                            x += 1

                                        if submitted:
                                            now = datetime.now()
                                            dt = now.strftime("%d/%m/%Y - %H:%M:%S")
                                            jobs[job]["Notes"][f'{dt} - {st.session_state["name"]}'] = text 
                                            self.write_to_json(jobs, job)
                                            x += 1

                                        notes = []
                                        for keys, values in jobs[job]["Notes"].items():
                                            notes.append(keys)
                                        
                                        notes.sort()
                                        notes.reverse()

                                        for note in notes:

                                            st.text_area(label=note, key=x, value=jobs[job]["Notes"][note], disabled=True)
                                            x += 1
                                        
                    except:
                        pass



        elif st.session_state["authentication_status"] is False:
            st.error('Username/password is incorrect')
            st.warning('Reset Password?')
        elif st.session_state["authentication_status"] is None:
            st.warning('Please enter your username and password')


        if authentication_status is False:
            try:
                if authenticator.reset_password(username, 'Reset password', 'main'):
                    st.success('Passord successfully changed!')
            except Exception as e:
                st.error(e)



App().main()